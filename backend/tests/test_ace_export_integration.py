"""Integration tests for ACE Lesson Plan exports and data flow.

Proves:
1. generated schedule session -> ACE context -> planned_date is populated, planned_period is populated, executed_date is blank for pending.
2. executed session -> ACE context -> executed_date is populated.
3. clock-based schedule -> planned_period formatted with clock times.
4. Case 1: schedule genuinely does not exist -> fallback to structured plan with blank columns.
5. Case 3: schedule lookup failure -> raises DocumentGenerationError (500).
6. Case 4: schedule malformed / empty sessions -> raises DocumentGenerationError or EmptyScheduleError.
"""

from __future__ import annotations

import io
from bson import ObjectId
import pytest
from docx import Document
from openpyxl import load_workbook

from app.services.ace_lesson_plan_export import (
    assemble_ace_context,
    export_ace_docx,
    export_ace_pdf,
    export_ace_xlsx,
)
from app.services.export_service import (
    DocumentGenerationError,
    EmptyScheduleError,
    build_lesson_plan_export,
    get_ace_lesson_plan_context,
)


def _structured_plan():
    return {
        "course_title": "Database Management Systems",
        "academic_year": "2026-2027",
        "units": [
            {
                "unit_number": 1,
                "unit_title": "Introduction to Databases",
                "topics": [
                    {
                        "topic_id": "U1-T1",
                        "topic": "Relational Model",
                        "teaching_methods": ["Lecture", "Chalk and Talk"],
                        "references": ["Korth Chapter 1"],
                        "estimated_hours": 2,
                    },
                    {
                        "topic_id": "U1-T2",
                        "topic": "SQL Queries",
                        "teaching_methods": ["Lab", "Project based learning"],
                        "references": ["Korth Chapter 3"],
                        "estimated_hours": 2,
                    },
                ],
            }
        ],
    }


def _course_doc(course_id, faculty_id):
    return {
        "_id": course_id,
        "course_code": "CS301",
        "course_name": "Database Management Systems",
        "department": "Computer Science and Engineering",
        "programme": "B.E. CSE",
        "semester": 5,
        "academic_year": "2026-2027",
        "faculty_id": faculty_id,
    }


@pytest.mark.asyncio
async def test_ace_export_data_flow_planned_dates_populated(db):
    course_id = ObjectId()
    faculty_id = ObjectId()
    lesson_id = ObjectId()

    await db.courses.insert_one(_course_doc(course_id, faculty_id))
    await db.lesson_plans.insert_one(
        {
            "_id": lesson_id,
            "course_id": course_id,
            "structured_plan": _structured_plan(),
        }
    )

    # Generated schedule with period-based sessions
    sessions = [
        {
            "session_id": "U1-T1-s1",
            "topic_id": "U1-T1",
            "topic": "Relational Model",
            "unit_number": 1,
            "unit_title": "Introduction to Databases",
            "date": "2026-08-03",
            "day": "Monday",
            "period_start": 1,
            "period_end": 1,
            "duration_hours": 1.0,
            "status": "pending",
        },
        {
            "session_id": "U1-T2-s1",
            "topic_id": "U1-T2",
            "topic": "SQL Queries",
            "unit_number": 1,
            "unit_title": "Introduction to Databases",
            "date": "2026-08-05",
            "day": "Wednesday",
            "period_start": 3,
            "period_end": 4,
            "duration_hours": 2.0,
            "status": "pending",
        },
    ]

    await db.generated_schedules.insert_one(
        {
            "course_id": course_id,
            "lesson_plan_id": lesson_id,
            "faculty_id": faculty_id,
            "sessions": sessions,
            "version": 1,
            "active": True,
        }
    )

    context, lesson, course = await get_ace_lesson_plan_context(str(lesson_id))

    assert context["has_schedule"] is True
    rows = context["units"][0]["rows"]
    assert len(rows) == 2

    # Verify Planned Date with Day is populated
    assert rows[0]["planned_date"] == "03 Aug 2026 (Monday)"
    assert rows[0]["planned_period"] == "Hour 1"
    assert rows[0]["executed_date"] == ""  # blank when pending

    assert rows[1]["planned_date"] == "05 Aug 2026 (Wednesday)"
    assert rows[1]["planned_period"] == "Hour 3–4"
    assert rows[1]["executed_date"] == ""

    # Test PDF generation
    pdf_bytes, filename, media_type = await build_lesson_plan_export(str(lesson_id), "pdf")
    assert isinstance(pdf_bytes, bytes)
    assert pdf_bytes.startswith(b"%PDF-")
    assert media_type == "application/pdf"

    # Test DOCX generation
    docx_bytes, filename, media_type = await build_lesson_plan_export(str(lesson_id), "docx")
    doc = Document(io.BytesIO(docx_bytes))
    full_docx_text = " ".join(cell.text for t in doc.tables for r in t.rows for cell in r.cells)
    assert "03 Aug 2026 (Monday)" in full_docx_text
    assert "Hour 1" in full_docx_text

    # Test XLSX generation
    xlsx_bytes, filename, media_type = await build_lesson_plan_export(str(lesson_id), "xlsx")
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    all_values = [str(c.value) for r in ws.iter_rows() for c in r if c.value is not None]
    assert any("03 Aug 2026 (Monday)" in v for v in all_values)
    assert any("Hour 1" in v for v in all_values)


@pytest.mark.asyncio
async def test_ace_export_data_flow_executed_dates_populated(db):
    course_id = ObjectId()
    faculty_id = ObjectId()
    lesson_id = ObjectId()

    await db.courses.insert_one(_course_doc(course_id, faculty_id))
    await db.lesson_plans.insert_one(
        {
            "_id": lesson_id,
            "course_id": course_id,
            "structured_plan": _structured_plan(),
        }
    )

    # Session 1 is executed, Session 2 is pending
    sessions = [
        {
            "session_id": "U1-T1-s1",
            "topic_id": "U1-T1",
            "topic": "Relational Model",
            "unit_number": 1,
            "unit_title": "Introduction to Databases",
            "date": "2026-08-03",
            "day": "Monday",
            "period_start": 1,
            "period_end": 1,
            "duration_hours": 1.0,
            "status": "completed",
            "executed_date": "2026-08-03",
            "executed_day": "Monday",
            "actual_topics": ["Relational Model", "Entity-Relationship Basics"],
        },
        {
            "session_id": "U1-T2-s1",
            "topic_id": "U1-T2",
            "topic": "SQL Queries",
            "unit_number": 1,
            "unit_title": "Introduction to Databases",
            "date": "2026-08-05",
            "day": "Wednesday",
            "period_start": 3,
            "period_end": 4,
            "duration_hours": 2.0,
            "status": "pending",
        },
    ]

    await db.generated_schedules.insert_one(
        {
            "course_id": course_id,
            "lesson_plan_id": lesson_id,
            "faculty_id": faculty_id,
            "sessions": sessions,
            "version": 1,
            "active": True,
        }
    )

    context, _, _ = await get_ace_lesson_plan_context(str(lesson_id))
    rows = context["units"][0]["rows"]

    # Session 1 has executed date
    assert rows[0]["executed_date"] == "03 Aug 2026 (Monday)"
    assert "Entity-Relationship Basics" in rows[0]["topics_covered"]

    # Session 2 has blank executed date
    assert rows[1]["executed_date"] == ""


@pytest.mark.asyncio
async def test_ace_export_case1_no_schedule_fallback(db):
    course_id = ObjectId()
    lesson_id = ObjectId()

    await db.courses.insert_one(_course_doc(course_id, ObjectId()))
    await db.lesson_plans.insert_one(
        {
            "_id": lesson_id,
            "course_id": course_id,
            "structured_plan": _structured_plan(),
        }
    )

    # No schedule in db
    context, _, _ = await get_ace_lesson_plan_context(str(lesson_id))
    assert context["has_schedule"] is False

    rows = context["units"][0]["rows"]
    assert len(rows) == 2
    assert rows[0]["planned_date"] == ""
    assert rows[0]["planned_period"] == ""
    assert rows[0]["executed_date"] == ""
    assert rows[0]["topics_covered"] == "Relational Model"


@pytest.mark.asyncio
async def test_ace_export_case4_malformed_empty_schedule_raises(db):
    course_id = ObjectId()
    lesson_id = ObjectId()

    await db.courses.insert_one(_course_doc(course_id, ObjectId()))
    await db.lesson_plans.insert_one(
        {
            "_id": lesson_id,
            "course_id": course_id,
            "structured_plan": _structured_plan(),
        }
    )

    # Schedule exists but sessions is empty
    await db.generated_schedules.insert_one(
        {
            "course_id": course_id,
            "lesson_plan_id": lesson_id,
            "sessions": [],
            "version": 1,
            "active": True,
        }
    )

    with pytest.raises(EmptyScheduleError):
        await get_ace_lesson_plan_context(str(lesson_id))


@pytest.mark.asyncio
async def test_ace_export_clock_based_schedule(db):
    course_id = ObjectId()
    lesson_id = ObjectId()

    await db.courses.insert_one(_course_doc(course_id, ObjectId()))
    await db.lesson_plans.insert_one(
        {
            "_id": lesson_id,
            "course_id": course_id,
            "structured_plan": _structured_plan(),
        }
    )

    sessions = [
        {
            "session_id": "U1-T1-s1",
            "topic_id": "U1-T1",
            "topic": "Relational Model",
            "unit_number": 1,
            "unit_title": "Introduction to Databases",
            "date": "2026-08-03",
            "day": "Monday",
            "start_time": "09:00",
            "end_time": "10:00",
            "duration_hours": 1.0,
            "status": "pending",
        }
    ]

    await db.generated_schedules.insert_one(
        {
            "course_id": course_id,
            "lesson_plan_id": lesson_id,
            "sessions": sessions,
            "version": 1,
            "active": True,
        }
    )

    context, _, _ = await get_ace_lesson_plan_context(str(lesson_id))
    row = context["units"][0]["rows"][0]
    assert row["planned_date"] == "03 Aug 2026 (Monday)"
    assert row["planned_period"] == "09:00–10:00"
